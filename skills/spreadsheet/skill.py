"""Spreadsheet skill for CSV and Excel file operations."""

import asyncio
import csv
from typing import Any
from datetime import datetime, timedelta, timezone
from uuid import uuid4

from core.approval_gate import ApprovalGate, ApprovalRequest, ApprovalActionType
from core.observability import (
    TraceEventType,
    TraceComponent,
    TraceLevel,
    TraceEvent,
    TraceEmitter,
    MemoryTraceEmitter,
)


class SpreadsheetSkill:
    """Spreadsheet skill for CSV and Excel file operations."""

    def __init__(
        self,
        emitter: TraceEmitter | None = None,
        approval_gate: ApprovalGate | None = None,
    ) -> None:
        """Initialize the spreadsheet skill.

        Args:
            emitter: Trace emitter for observability
            approval_gate: Approval gate for write operations
        """
        self._emitter = emitter or MemoryTraceEmitter()
        self._approval_gate = approval_gate

    async def read_csv(self, path: str) -> list[dict[str, Any]]:
        """Read CSV file, returns list of row dicts (first row as headers).

        Args:
            path: Path to the CSV file

        Returns:
            List of row dicts with headers as keys

        Raises:
            FileNotFoundError: If file does not exist
        """
        start_time = asyncio.get_event_loop().time()

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet read_csv",
                data={"operation": "read_csv", "path": path},
                duration_ms=0,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        # Check if file exists - raise FileNotFoundError outside try-except
        import os
        if not os.path.exists(path):
            raise FileNotFoundError(f"CSV file not found: {path}")

        rows = []
        with open(path, "r", newline="", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                rows.append(dict(row))

        duration_ms = int((asyncio.get_event_loop().time() - start_time) * 1000)

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet read_csv completed",
                data={"operation": "read_csv", "path": path, "rows_read": len(rows)},
                duration_ms=duration_ms,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        return rows

    async def write_csv(self, path: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
        """Write list of dicts to CSV.

        Args:
            path: Path to the CSV file
            rows: List of dicts to write

        Returns:
            Dict with success and rows_written

        Raises:
            FileNotFoundError: If file does not exist
        """
        if self._approval_gate:
            request = ApprovalRequest(
                request_id=str(uuid4()),
                task_id=str(uuid4()),
                session_id="default",
                action_type=ApprovalActionType.FILE_WRITE,
                action_description="spreadsheet write_csv",
                action_parameters={"path": path},
                risk_level="low",
                reason_for_approval="Spreadsheet write requires approval per policy",
                expires_at=datetime.now(timezone.utc) + timedelta(seconds=300),
            )
            response = await self._approval_gate.request_approval(request)
            approved = response.approved
            if not approved:
                return {
                    "success": False,
                    "rows_written": 0,
                }

        start_time = asyncio.get_event_loop().time()

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet write_csv",
                data={"operation": "write_csv", "path": path},
                duration_ms=0,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        if not rows:
            rows_written = 0
        else:
            with open(path, "w", newline="", encoding="utf-8") as csvfile:
                fieldnames = rows[0].keys()
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
                rows_written = len(rows)

        duration_ms = int((asyncio.get_event_loop().time() - start_time) * 1000)

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet write_csv completed",
                data={"operation": "write_csv", "path": path, "rows_written": rows_written},
                duration_ms=duration_ms,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        return {
            "success": True,
            "rows_written": rows_written,
        }

    async def read_excel(self, path: str, sheet: str | None = None) -> list[dict[str, Any]]:
        """Read first sheet (or named sheet) from .xlsx file.

        Args:
            path: Path to the Excel file
            sheet: Optional sheet name to read (reads first sheet if None)

        Returns:
            List of row dicts

        Raises:
            FileNotFoundError: If file does not exist
        """
        start_time = asyncio.get_event_loop().time()

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet read_excel",
                data={"operation": "read_excel", "path": path, "sheet": sheet},
                duration_ms=0,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        # Check if file exists - raise FileNotFoundError outside try-except
        import os
        if not os.path.exists(path):
            raise FileNotFoundError(f"Excel file not found: {path}")

        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        if sheet:
            ws = wb[sheet]
        else:
            ws = wb.active

        rows = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(cell) if cell is not None else "" for cell in row]
            else:
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                rows.append(row_dict)

        wb.close()

        duration_ms = int((asyncio.get_event_loop().time() - start_time) * 1000)

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet read_excel completed",
                data={"operation": "read_excel", "path": path, "rows_read": len(rows)},
                duration_ms=duration_ms,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        return rows

    async def write_excel(self, path: str, rows: list[dict[str, Any]], sheet: str = "Sheet1") -> dict[str, Any]:
        """Write list of dicts to .xlsx file.

        Args:
            path: Path to the Excel file
            rows: List of dicts to write
            sheet: Sheet name (default "Sheet1")

        Returns:
            Dict with success and rows_written
        """
        if self._approval_gate:
            request = ApprovalRequest(
                request_id=str(uuid4()),
                task_id=str(uuid4()),
                session_id="default",
                action_type=ApprovalActionType.FILE_WRITE,
                action_description="spreadsheet write_excel",
                action_parameters={"path": path},
                risk_level="low",
                reason_for_approval="Spreadsheet write requires approval per policy",
                expires_at=datetime.now(timezone.utc) + timedelta(seconds=300),
            )
            response = await self._approval_gate.request_approval(request)
            approved = response.approved
            if not approved:
                return {
                    "success": False,
                    "rows_written": 0,
                }

        start_time = asyncio.get_event_loop().time()

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet write_excel",
                data={"operation": "write_excel", "path": path, "sheet": sheet},
                duration_ms=0,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = sheet

        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h, "") for h in headers])
            rows_written = len(rows)
        else:
            rows_written = 0

        wb.save(path)

        duration_ms = int((asyncio.get_event_loop().time() - start_time) * 1000)

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet write_excel completed",
                data={"operation": "write_excel", "path": path, "rows_written": rows_written},
                duration_ms=duration_ms,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        return {
            "success": True,
            "rows_written": rows_written,
        }

    async def sheet_names(self, path: str) -> list[str]:
        """Return list of sheet names in an .xlsx file.

        Args:
            path: Path to the Excel file

        Returns:
            List of sheet names

        Raises:
            FileNotFoundError: If file does not exist
        """
        start_time = asyncio.get_event_loop().time()

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet sheet_names",
                data={"operation": "sheet_names", "path": path},
                duration_ms=0,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        # Check if file exists - raise FileNotFoundError outside try-except
        import os
        if not os.path.exists(path):
            raise FileNotFoundError(f"Excel file not found: {path}")

        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()

        duration_ms = int((asyncio.get_event_loop().time() - start_time) * 1000)

        try:
            event = TraceEvent(
                event_type=TraceEventType.SPREADSHEET_OPERATION,
                component=TraceComponent.SPREADSHEET_SKILL,
                level=TraceLevel.INFO,
                message="Spreadsheet sheet_names completed",
                data={"operation": "sheet_names", "path": path, "sheet_count": len(names)},
                duration_ms=duration_ms,
            )
            await self._emitter.emit(event)
        except Exception:
            # Trace emission failure - non-critical, continue
            pass

        return names
